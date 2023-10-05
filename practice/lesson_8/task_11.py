"""
Прочитайте файл из прошлого задания и выведите данные в формате "{Фамилия} {Имя} {Возраст}".
"""

import openpyxl

wb = openpyxl.load_workbook('file_10.xlsx')
worksheet = wb.active

for row in range(worksheet.max_row):
    for column in worksheet.iter_cols():
        value = column[row].value
        print(str(value).ljust(15), end='')
    print('')

# print([sheet["A1"].value, sheet["B1"].value, sheet["C1"].value])
# print([sheet["A2"].value, sheet["B2"].value, sheet["C2"].value])
